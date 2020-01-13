from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import load_workbook

# prime ids: non-negative whole numbers and no duplicates
# prices and payments: non-negative numbers
# quantities, non-prime ids: non-negative whole numbers
# customer, product names: less than 20 character strings
# customer emails: <50 char string
# date: valid date format
# active: 0 (not active) or 1 (active)

dv_non_neg_whole = DataValidation(type='whole',
                                  operator='greaterThanOrEqual',
                                  formula1=0)
dv_date = DataValidation(type='date')
dv_string_len20 = DataValidation(type="textLength",
                                 operator="lessThanOrEqual",
                                 formula1=20)
dv_string_len50 = DataValidation(type="textLength",
                                 operator="lessThanOrEqual",
                                 formula1=50)
dv_0_or_1 = DataValidation(type='whole',
                           operator='between',
                           formula1=0,
                           formula2=1)
dv_non_neg = DataValidation(type='decimal',
                            operator='greaterThanOrEqual',
                            formula1=0)

dvs = [
    dv_0_or_1, dv_date, dv_non_neg_whole, dv_string_len50, dv_string_len20,
    dv_non_neg
]

prime_ids = {'invoices': 'i_id', 'customers': 'c_id', 'products': 'p_id'}
non_prime_ids = {
    'invoices': 'c_id',
    'invoice_details': 'i_id, p_id',
    'customers': '',
    'products': ''
}

wb = load_workbook('sample_db_1.xlsx')

for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    for col in ws[1]:
        header = col.value
        if header.find('id') != -1:
            if non_prime_ids.get(col.parent.title).find(col.value) != -1:
                # print(col.parent.title, col.value, col.column_letter)
                dv_non_neg_whole.add(
                    f'{col.column_letter}2:{col.column_letter}1048576')
                ws.add_data_validation(dv_non_neg_whole)
            else:
                # print("prime", col.parent.title, col.value, col.column_letter)
                no_dup_range = f'=countif({col.column_letter}:{col.column_letter}, {col.column_letter}2)<=1'
                no_dup_cells = f'{col.column_letter}2:{col.column_letter}1048576'
                dv_no_dup = DataValidation(type='custom',
                                           formula1=no_dup_range)
                dv_no_dup.add(no_dup_cells)
                ws.add_data_validation(dv_no_dup)
                dvs.append(dv_no_dup)
        elif header.find('date') != -1:
            dv_date.add(f'{col.column_letter}2:{col.column_letter}1048576')
            ws.add_data_validation(dv_date)
        elif header.find('name') != -1:
            dv_string_len20.add(
                f'{col.column_letter}2:{col.column_letter}1048576')
            ws.add_data_validation(dv_string_len20)
        elif header.find('email') != -1:
            dv_string_len50.add(
                f'{col.column_letter}2:{col.column_letter}1048576')
            ws.add_data_validation(dv_string_len50)
        elif header.find('quant') != -1:
            dv_non_neg_whole.add(
                f'{col.column_letter}2:{col.column_letter}1048576')
            ws.add_data_validation(dv_non_neg_whole)
        elif header.find('payment') != -1 or header.find('price') != -1:
            dv_non_neg.add(f'{col.column_letter}2:{col.column_letter}1048576')
            ws.add_data_validation(dv_non_neg)
        elif header.find('active') != -1:
            dv_0_or_1.add(f'{col.column_letter}2:{col.column_letter}1048576')
            ws.add_data_validation(dv_0_or_1)

for dv in dvs:
    dv.errorTitle = 'Invalid Entry'
    if type(dv.formula1) == str and dv.formula1.find('countif') != -1:
        dv.error = 'Duplicate entry'
    else:
        dv.error = f'Need to enter data of type: {dv.type}. ' + f'And its value needs to: {dv.operator} {dv.formula1}' + (
            f' and {dv.formula2}' if dv.formula2 != None else '')

wb.save('sample_db_1.xlsx')