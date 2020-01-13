from openpyxl import load_workbook


def get_col_let_by_header(header, ws):
    for h in ws[1]:
        if h.value == header:
            return h.column_letter
    return None


def get_header_list(ws):
    return [h.value for h in ws[1]]


def retrieve_info_from_id(id, id_type, ws):
    i_id_col = get_col_let_by_header(id_type, ws)
    results = []
    for cell in ws[i_id_col]:
        if cell.value == id:
            results.append({
                h: c.value
                for h, c in zip(get_header_list(ws), ws[cell.row])
            })
    return results


def print_invoice(re):
    # {'date': datetime.datetime(2019, 3, 2, 0, 0), 'f_name': 'Allen', 'l_name': 'Butterfield', 'products': [{'product_name': 'Cardio II', 'quantity': 1, 'price': 500}, {'product_name': 'Strength Training II', 'quantity': 1, 'price': 500}]}
    string = f'-------------------------------------  Invoice -------------------------------------\n\n'
    string += 'First Name'.ljust(15) + 'Last Name'.ljust(15) + 'Date'.ljust(
        15) + 'Amount ($)'.ljust(15) + 'Product'.ljust(25) + 'Quantity'.ljust(
            15) + '\n'
    for item in re.get('products'):
        for i in content:
            string += str(re.get(i)).replace('00:00:00', '').ljust(15)
        for i in sub_content:
            string += str(
                item.get(i)).ljust(15) if i != 'product_name' else str(
                    item.get(i)).ljust(25)
        string += '\n'
    return string + '\n\n\n'


content = ['f_name', 'l_name', 'date']
sub_content = ['price', 'product_name', 'quantity']


def generate_invoice(iid, db='merged_sample_db.xlsx'):
    wb = load_workbook(db)
    result = {}
    invoice_info = retrieve_info_from_id(iid, 'i_id', wb['invoices'])[0]
    result['date'] = invoice_info.get('date')
    cid = invoice_info.get('c_id')
    customer_info = retrieve_info_from_id(cid, 'c_id', wb['customers'])[0]
    result['f_name'] = customer_info.get('f_name')
    result['l_name'] = customer_info.get('l_name')
    result['products'] = []
    quantity_info = retrieve_info_from_id(iid, 'i_id', wb['invoice_details'])
    # [{'i_id': 4, 'p_id': 2, 'quantity': 1}, {'i_id': 4, 'p_id': 4, 'quantity': 1}]
    for item in quantity_info:
        product_info = retrieve_info_from_id(item.get('p_id'), 'p_id',
                                             wb['products'])[0]
        product_name = product_info.get('name')
        price = product_info.get('unit_price')
        result['products'].append({
            'product_name': product_name,
            'quantity': item.get('quantity'),
            'price': price
        })
    print(result)
    return print_invoice(result)


try:
    invoice_report = open('invoice_report.txt', 'w')
    invoice_report.write(
        generate_invoice(4) + generate_invoice(26) + generate_invoice(10))
finally:
    invoice_report.close()
print(generate_invoice(4))

# -------------------------------------  Invoice -------------------------------------

# First Name     Last Name      Date           Amount ($)     Product                  Quantity
# Allen          Butterfield    2019-03-02     500            Cardio II                1
# Allen          Butterfield    2019-03-02     500            Strength Training II     1