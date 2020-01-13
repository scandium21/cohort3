from openpyxl import load_workbook


class DiffSheetNums(Exception):
    def __init__(self, input, target):
        message = f"{input} does not contain same number of sheet as {target}"
        super().__init__(message)


class DiffSheetNames(Exception):
    def __init__(self, input, target):
        message = f"{input} has sheet whose names are different from those in {target}"
        super().__init__(message)


def merge_files(file_des, main_file, *rest_files_to_merge):
    wb1 = load_workbook(main_file)
    for f in rest_files_to_merge:
        wb_to_merge = load_workbook(f)
        if len(wb1.sheetnames) != len(wb_to_merge.sheetnames):
            raise DiffSheetNums(main_file, f)
        elif wb1.sheetnames != wb_to_merge.sheetnames:
            raise DiffSheetNames(main_file, f)
        for name in wb1.sheetnames:
            # print("now accessing sheet: ", name)
            ws1 = wb1[name]
            ws_to_merge = wb_to_merge[name]
            for row_to_merge, row1 in zip(ws_to_merge.rows, ws1.rows):
                if row_to_merge[0].value == row1[0].value:
                    # print(row_to_merge[0], row_to_merge[0].value)
                    continue
                ws1.append((cell.value for cell in row_to_merge))
    wb1.save(file_des)


merge_files('merged_sample_db.xlsx', 'sample_db_1.xlsx', 'sample_db_2.xlsx')
