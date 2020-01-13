from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl import Workbook
import random
import datetime

print(get_column_letter(1), column_index_from_string('AC'))
# A 29

wb = load_workbook('sample_db_copy.xlsx')

print(wb.sheetnames)
# ['invoices', 'invoice_details', 'customers', 'products']

invoices = wb['invoices']
print(invoices.title, invoices.max_row, invoices.max_column)
# invoices 32 4
for header, val in zip(invoices['1'], invoices['2']):
    print(header.value, val.value, type(val.value))
# i_id 1 <class 'int'>
# c_id 1 <class 'int'>
# date 2019-03-01 00:00:00 <class 'datetime.datetime'>
# total_payment 400 <class 'int'>
print(invoices.cell(row=3, column=3).value)
# 2019-03-01 00:00:00

ws = wb['products']
print('***', list(ws.rows)[1:])

wb.create_sheet('learning_openpyxl', 0)
wb.save('openpyxl_ex.xlsx')

# set row, col height and width
# invoices.row_dimensions[1].height = 70
# invoices.column_dimensions['A'].width = 20

# change font
# invoices['A1'].font = Font(sz=14, bold=True, italic=True)

# create new wb
wb = Workbook()
sheet = wb.create_sheet('my_sheet', 0)
for i in range(1, 11):
    sheet[f'A{i}'].value = random.randint(1, 100)

wb.save('openpyxl_ex.xlsx')
